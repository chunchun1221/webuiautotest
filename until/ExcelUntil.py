import os.path
from openpyxl import load_workbook

from chaper99.until.LoggingConfig import Loging


class ExcelUtil(object):
    def __init__(self, excel_path=None, sheet_name=None):
        self.logger = Loging().get_log()
        if excel_path is None:
            current_path = os.path.abspath(os.path.dirname(__file__))
            excel_path = os.path.join(current_path, "../data/casedata.xlsx")
        self.excel_path = excel_path
        self.sheet_name = sheet_name or 'Sheet1'

        try:
            self.excel_data = load_workbook(self.excel_path)
            self.sheet = self.excel_data[self.sheet_name]
            self.logger.debug(f"Loaded workbook from {self.excel_path}")
        except Exception as e:
            self.logger.error(f"Failed to load workbook or sheet: {e}")
            raise  # 让调用者知道初始化失败了

    def get_all_data(self):
        rows = self.sheet.rows
        row_num = self.sheet.max_row
        column = self.sheet.max_column

        if row_num <= 1:
            self.logger.info("请注意，工作表中没有足够的测试数据。")
            return []

        all_case = [[cell.value for cell in row] for row in rows]
        return all_case


# if __name__ == "__main__":
#     sheet = "Login"
#     try:
#         file = ExcelUtil(sheet_name=sheet)
#         print(file.get_all_data())
#     except Exception as e:
#         print(f"遇到错误: {e}")
